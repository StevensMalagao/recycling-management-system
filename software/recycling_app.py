import serial
import serial.tools.list_ports
import time
import re
import openpyxl
import os
import customtkinter as ctk
from tkinter import messagebox
import threading
from PIL import Image, ImageDraw
import webbrowser

# --- Configuration Constants ---
EXCEL_LOG_FILE = "Recycling_Logs.xlsx"
COLLECTORS_DB_FILE = "Collectors_Database.xlsx"
BAUD_RATE = 115200
LOGO_PATH = "assets/logo.png" 
LOGO_SIZE = (400, 120)
REFRESH_ICON_PATH = "assets/refresh_icon.png"

# --- Data Structures ---
MATERIALS_BY_CATEGORY = {
    "Metals": [
        "Aluminum", "Scrap Metal", "Copper", "Bronze", 
        "Antimony", "Steel", "Other Metals"
    ],
    "Paper & Cardboard": [
        "Archive", "Cardboard", "Trays/Panels", "Newspaper", 
        "Foldable", "Tetra Pack", "Plasticized", "Kraft", 
        "Other Paper/Cardboard"
    ],
    "Plastics": [
        "Acrylic", "Paste", "PET", "PVC", "White Plastic", 
        "Polyethylene", "Blown Plastic", "Polypropylene", 
        "Other Plastics"
    ],
    "Glass": ["Other Glass"],
    "Textile": ["Other Textiles"],
    "Wood": ["Other Wood"]
}

PACKAGING_WEIGHTS = {
    "None": 0.0,
    "Burlap Bag": 1.5,
    "Balloon": 2.5,
    "Tarp": 2.0
}

# --- Utility Functions ---

def manage_collectors_db(operation, file_name, sheet_name="Collectors"):
    """
    Creates or reads the database of waste collectors (recyclers).
    """
    if not os.path.exists(file_name):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Collector_Name", "Collector_ID"])
            ws.append(["Example Collector 1", "123456789"])
            wb.save(file_name)
        except Exception as e:
            print(f"Error creating Excel database ('{file_name}'): {e}")
            return {"Error": "N/A"}

    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        
        if operation == "read":
            collectors_dict = {}
            for row in range(2, ws.max_row + 1):
                name_cell = ws.cell(row=row, column=1).value
                id_cell = ws.cell(row=row, column=2).value
                
                if name_cell:
                    name = str(name_cell).strip()
                    collector_id = str(id_cell).strip() if id_cell else "N/A"
                    collectors_dict[name] = collector_id
            
            return collectors_dict if collectors_dict else {"Empty List": "N/A"}

    except Exception as e:
        print(f"Error reading Excel ('{file_name}'): {e}")
        return {"Error": "N/A"}

def append_data_to_excel(file_name, entry_data):
    """
    Saves the transaction to Excel. Creates a new sheet per month 
    and dynamically adds columns for materials.
    """
    date_val = entry_data["date"]
    collector = entry_data["collector"]
    collector_id = entry_data["collector_id"]
    material = entry_data["material"]
    net_weight = float(entry_data["net_weight"])

    months_en = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    
    # Determine Sheet Name (Month-Year)
    try:
        parts = date_val.split('/')
        if len(parts) == 3:
            day, month_num, year = parts
            sheet_title = f"{months_en[int(month_num) - 1]}-{year}"
        else:
            raise ValueError("Invalid date format")
    except Exception as e:
        print(f"Date parse warning: {e}. Using 'General_Logs'.")
        sheet_title = "General_Logs"
    
    # Standard Headers
    HEADERS = ["DATE", "NAME", "ID", "TOTAL"]

    try:
        if os.path.exists(file_name):
            wb = openpyxl.load_workbook(file_name)
            ws = wb[sheet_title] if sheet_title in wb.sheetnames else wb.create_sheet(sheet_title)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_title

        # Initialize Headers if new sheet
        if ws.cell(row=1, column=1).value is None:
            for idx, header in enumerate(HEADERS):
                ws.cell(row=1, column=idx + 1).value = header

        headers_list = [cell.value for cell in ws[1]]
        
        # Helper to find or create column
        def get_col_idx(col_name):
            nonlocal headers_list
            try:
                return headers_list.index(col_name) + 1
            except ValueError:
                # Insert new material before 'TOTAL' or at the end
                try:
                    total_idx = headers_list.index("TOTAL")
                except ValueError:
                    total_idx = len(headers_list)
                    ws.cell(row=1, column=total_idx + 1).value = "TOTAL"
                    headers_list.append("TOTAL")

                ws.insert_cols(total_idx + 1)
                ws.cell(row=1, column=total_idx + 1).value = col_name
                headers_list.insert(total_idx, col_name)
                return total_idx + 1

        date_idx = get_col_idx("DATE")
        name_idx = get_col_idx("NAME")
        id_idx = get_col_idx("ID")
        mat_idx = get_col_idx(material)
        total_idx = get_col_idx("TOTAL")

        # Check if row exists for this collector on this date
        target_row = -1
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=date_idx).value == date_val and \
               ws.cell(row=r, column=name_idx).value == collector:
                target_row = r
                break
        
        # Create new row if not found
        if target_row == -1:
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=date_idx).value = date_val
            ws.cell(row=target_row, column=name_idx).value = collector
            ws.cell(row=target_row, column=id_idx).value = collector_id
            
            # Initialize other numeric columns to 0
            for c in range(1, len(headers_list) + 1):
                if c not in [date_idx, name_idx, id_idx]:
                    if ws.cell(row=target_row, column=c).value is None:
                        ws.cell(row=target_row, column=c).value = 0.0

        # Update values
        current_mat_val = float(ws.cell(row=target_row, column=mat_idx).value or 0.0)
        ws.cell(row=target_row, column=mat_idx).value = current_mat_val + net_weight
        
        current_total_val = float(ws.cell(row=target_row, column=total_idx).value or 0.0)
        ws.cell(row=target_row, column=total_idx).value = current_total_val + net_weight

        wb.save(file_name)
        return True

    except Exception as e:
        print(f"Excel Write Error: {e}")
        messagebox.showerror("Save Error", f"Could not save to '{sheet_title}'.\nError: {e}\nEnsure file is closed.")
        return False

def handle_serial_error(gui_app, port):
    if gui_app.running:
        messagebox.showerror("Connection Error", f"Lost connection to {port}.\nApplication will close.")
        if gui_app.winfo_exists():
            gui_app.destroy()

def read_serial_data(port, baud_rate, gui_app):
    ser = None
    try:
        ser = serial.Serial(port, baud_rate, timeout=1)
        print(f"Opening serial port {port} at {baud_rate} baud...")
        time.sleep(2)
        while gui_app.running:
            if ser.in_waiting > 0:
                line = ser.readline().decode('utf-8').strip()
                if line: gui_app.process_serial_line(line)
            time.sleep(0.1)
    except serial.SerialException:
        print(f"Connection failed on {port}.")
        if gui_app.running:
            gui_app.after(0, handle_serial_error, gui_app, port)
    except Exception as e:
        print(f"Unexpected serial thread error: {e}")
    finally:
        if ser and ser.is_open: ser.close()
        print("Serial port closed.")

# --- Main GUI Class ---

class RecyclingApp(ctk.CTk):
    def __init__(self, selected_port):
        super().__init__()
        
        self.weight_lock = threading.Lock()
        self.serial_port = selected_port
        self.current_gross_weight = 0.0
        self.running = True
        
        self.collectors_db = manage_collectors_db("read", COLLECTORS_DB_FILE)
        collector_names = sorted(list(self.collectors_db.keys()))

        self.title(f"Asorate Registration System - {self.serial_port}")
        self.geometry("800x650") 
        self.resizable(True, True)
        self.grid_columnconfigure(0, weight=1)
        
        # Variables
        self.date_var = ctk.StringVar(value="Waiting...")
        self.gross_weight_var = ctk.StringVar(value="0.00 kg")
        self.tare_var = ctk.StringVar(value="0.00 kg")
        self.net_weight_var = ctk.StringVar(value="0.00 kg")
        self.selected_packaging = ctk.StringVar(value="None")
        self.selected_category = ctk.StringVar(value="Select Category")
        self.selected_material = ctk.StringVar(value="Select Material")
        self.selected_collector = ctk.StringVar(value="Select Collector")
        
        # Icons
        self.refresh_icon = None
        if os.path.exists(REFRESH_ICON_PATH):
            try:
                self.refresh_icon = ctk.CTkImage(Image.open(REFRESH_ICON_PATH), size=(20, 20))
            except Exception as e:
                print(f"Error loading icon: {e}")
        else:
             print(f"Warning: Icon '{REFRESH_ICON_PATH}' not found.")

        self.create_widgets(collector_names)
        self.start_serial_thread()
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def create_widgets(self, collector_list):
        # Logo
        if os.path.exists(LOGO_PATH):
            try:
                ctk.CTkLabel(self, image=ctk.CTkImage(Image.open(LOGO_PATH), size=LOGO_SIZE), text="").grid(row=0, column=0, padx=20, pady=(10, 5), sticky="n")
            except Exception as e: print(f"Error loading logo: {e}")
        
        # Data Display Frame
        data_frame = ctk.CTkFrame(self)
        data_frame.grid(row=1, column=0, padx=20, pady=15, sticky="ew")
        data_frame.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(data_frame, text="Date:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        ctk.CTkLabel(data_frame, textvariable=self.date_var).grid(row=0, column=1, padx=10, pady=5, sticky="w")
        
        ctk.CTkLabel(data_frame, text="Gross Weight:", font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, padx=10, pady=5, sticky="w")
        ctk.CTkLabel(data_frame, textvariable=self.gross_weight_var).grid(row=2, column=1, padx=10, pady=5, sticky="w")
        
        ctk.CTkLabel(data_frame, text="Tare (Packaging):", font=ctk.CTkFont(weight="bold")).grid(row=3, column=0, padx=10, pady=5, sticky="w")
        ctk.CTkLabel(data_frame, textvariable=self.tare_var).grid(row=3, column=1, padx=10, pady=5, sticky="w")
        
        ctk.CTkLabel(data_frame, text="Net Weight:", font=ctk.CTkFont(weight="bold")).grid(row=4, column=0, padx=10, pady=10, sticky="w")
        ctk.CTkLabel(data_frame, textvariable=self.net_weight_var).grid(row=4, column=1, padx=10, pady=10, sticky="w")

        # Input Frame
        input_frame = ctk.CTkFrame(self)
        input_frame.grid(row=2, column=0, padx=20, pady=25, sticky="nsew")
        input_frame.grid_columnconfigure((1, 3), weight=1)

        # Materials
        ctk.CTkLabel(input_frame, text="Category:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        ctk.CTkComboBox(input_frame, variable=self.selected_category, values=list(MATERIALS_BY_CATEGORY.keys()), state="readonly", command=self.update_material_list).grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        
        ctk.CTkLabel(input_frame, text="Material:").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.material_menu = ctk.CTkComboBox(input_frame, variable=self.selected_material, values=[], state="disabled")
        self.material_menu.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        # Packaging
        ctk.CTkLabel(input_frame, text="Packaging:").grid(row=1, column=2, padx=10, pady=10, sticky="w")
        ctk.CTkComboBox(input_frame, variable=self.selected_packaging, values=list(PACKAGING_WEIGHTS.keys()), state="readonly", command=self.update_weights).grid(row=1, column=3, padx=10, pady=10, sticky="ew")

        # Collector Selection
        ctk.CTkLabel(input_frame, text="Collector:").grid(row=0, column=2, padx=(20, 10), pady=10, sticky="w")
        collector_frame = ctk.CTkFrame(input_frame, fg_color="transparent")
        collector_frame.grid(row=0, column=3, padx=10, pady=10, sticky="ew")
        collector_frame.grid_columnconfigure(0, weight=1)
        
        self.collector_menu = ctk.CTkComboBox(collector_frame, variable=self.selected_collector, values=collector_list, state="readonly")
        self.collector_menu.grid(row=0, column=0, sticky="ew")
        
        # Refresh Button
        refresh_text = "" if self.refresh_icon else "Refresh"
        refresh_width = 35 if self.refresh_icon else 80
        ctk.CTkButton(
            collector_frame, 
            text=refresh_text, 
            image=self.refresh_icon, 
            width=refresh_width,
            command=self.refresh_collectors,
            fg_color="gray", 
            hover_color="darkgray"
        ).grid(row=0, column=1, padx=(10,0))
        
        # Main Buttons
        ctk.CTkButton(self, text="Register Transaction", command=self.save_data, font=ctk.CTkFont(weight="bold"), height=40, fg_color="#4EB9D7", hover_color="#357D92").grid(row=3, column=0, padx=20, pady=(10, 5), sticky="ew")
        ctk.CTkButton(self, text="Open Log File", command=self.open_log_file, font=ctk.CTkFont(weight="bold"), height=40, fg_color="#4ED761", hover_color="#3BA249").grid(row=4, column=0, padx=20, pady=(5, 5), sticky="ew")
        ctk.CTkButton(self, text="Manage Database", command=self.open_db_file, font=ctk.CTkFont(weight="bold"), height=40, fg_color="#46bb58", hover_color="#3BA249").grid(row=5, column=0, padx=20, pady=(5, 20), sticky="ew")

    def open_db_file(self):
        manage_collectors_db("leer", COLLECTORS_DB_FILE)
        try:
            webbrowser.open(os.path.abspath(COLLECTORS_DB_FILE))
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file:\n{e}")

    def refresh_collectors(self):
        print("Refreshing collector list...")
        self.collectors_db = manage_collectors_db("read", COLLECTORS_DB_FILE)
        new_list = sorted(list(self.collectors_db.keys()))
        
        self.collector_menu.configure(values=new_list)
        self.selected_collector.set("Select Collector")
        messagebox.showinfo("Updated", "Collector list refreshed from Excel.")

    def update_material_list(self, category):
        materials = MATERIALS_BY_CATEGORY.get(category, [])
        self.material_menu.configure(values=materials, state="readonly" if materials else "disabled")
        self.selected_material.set("Select Material" if materials else "No materials")

    def start_serial_thread(self):
        self.serial_thread = threading.Thread(target=read_serial_data, args=(self.serial_port, BAUD_RATE, self), daemon=True)
        self.serial_thread.start()
    
    def process_serial_line(self, line):
        # Matches: "Date: 12/12/2025, Weight: 10.50" 
        full_match = re.match(r'Date: ([\d/]+)(?:, Weight: ([\d.]+))?', line)
        
        if full_match:
            date_str, weight_str = full_match.groups()
            weight = float(weight_str) if weight_str else self.current_gross_weight
            self.update_ui_data(date_str, weight)
        elif re.match(r'^([\d.]+)$', line): #if line has separate weight only 
            self.update_ui_data(self.date_var.get(), float(line))

    def update_ui_data(self, date_str, gross_weight):
        with self.weight_lock:
            self.current_gross_weight = float(gross_weight)
        
        self.after(0, self._update_gui_labels, date_str)

    def _update_gui_labels(self, date_str):
        if date_str != "Waiting...": self.date_var.set(date_str)
        self.update_weights()
        
    def update_weights(self, _=None):
        tare = PACKAGING_WEIGHTS.get(self.selected_packaging.get(), 0.0)
        
        with self.weight_lock:
            current_gross = self.current_gross_weight
            
        self.gross_weight_var.set(f"{current_gross:.2f} kg")
        self.tare_var.set(f"{tare:.2f} kg")
        self.net_weight_var.set(f"{current_gross - tare:.2f} kg")

    def save_data(self):
        net_str = self.net_weight_var.get().replace(" kg", "")
        collector_name = self.selected_collector.get()
        
        with self.weight_lock:
            current_gross = self.current_gross_weight
        
        collector_id = self.collectors_db.get(collector_name, "N/A")
            
        data = {
            "date": self.date_var.get(),
            "gross_weight": current_gross,
            "packaging": self.selected_packaging.get(),
            "tare": PACKAGING_WEIGHTS.get(self.selected_packaging.get(), 0.0),
            "net_weight": float(net_str),
            "category": self.selected_category.get(),
            "material": self.selected_material.get().upper(),
            "collector": collector_name.upper(),
            "collector_id": collector_id, 
        }
        
        # Validation
        if any(v in ("Waiting...", "N/A") for v in [data["date"]]) or \
           any("Select" in v for v in [data["category"], data["material"], data["collector"]]):
            messagebox.showwarning("Incomplete Data", "Please fill all fields and wait for scale data."); return
        
        if data["net_weight"] <= 0:
            if not messagebox.askyesno("Zero/Negative Weight", f"Net weight is {data['net_weight']:.2f} kg. Register anyway?"): return

        if append_data_to_excel(EXCEL_LOG_FILE, data):
            messagebox.showinfo("Success", "Transaction saved successfully.")
            self.reset_fields()

    def reset_fields(self):
        self.selected_category.set("Select Category")
        self.material_menu.configure(values=[], state="disabled")
        self.selected_material.set("Select Material")
        self.selected_collector.set("Select Collector")
        self.selected_packaging.set("None")
        self.date_var.set("Waiting...")
        
        with self.weight_lock:
            self.current_gross_weight = 0.0
        self.update_weights()

    def open_log_file(self):
        if not os.path.exists(EXCEL_LOG_FILE): 
            messagebox.showwarning("Not Found", f"'{EXCEL_LOG_FILE}' does not exist yet."); return
        try: 
            webbrowser.open(os.path.abspath(EXCEL_LOG_FILE))
        except Exception as e: 
            messagebox.showerror("Error", f"Could not open file: {e}")

    def on_closing(self):
        if messagebox.askokcancel("Exit", "Do you want to close the application?"):
            self.running = False
            if hasattr(self, 'serial_thread') and self.serial_thread.is_alive(): 
                self.serial_thread.join(1)
            self.destroy()

# --- Connection Window ---

class COMPortSelection(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Connect Device")
        self.geometry("300x180")
        self.resizable(False, False)
        self.grid_columnconfigure(0, weight=1)

        self.refresh_icon = None
        if os.path.exists(REFRESH_ICON_PATH):
            try:
                self.refresh_icon = ctk.CTkImage(Image.open(REFRESH_ICON_PATH), size=(20, 20))
            except Exception: pass
        
        ctk.CTkLabel(self, text="Select Device Port:", font=ctk.CTkFont(size=14)).pack(padx=20, pady=(20, 5))
        
        com_frame = ctk.CTkFrame(self, fg_color="transparent")
        com_frame.pack(padx=20, pady=5, fill="x")
        com_frame.grid_columnconfigure(0, weight=1)

        self.selected_port = ctk.StringVar()
        self.port_combo = ctk.CTkComboBox(com_frame, variable=self.selected_port, state="readonly")
        self.port_combo.grid(row=0, column=0, sticky="ew")

        refresh_text = "" if self.refresh_icon else "Refresh"
        refresh_width = 35 if self.refresh_icon else 80
        
        ctk.CTkButton(
            com_frame, 
            text=refresh_text, 
            image=self.refresh_icon, 
            command=self.refresh_ports, 
            width=refresh_width,
            fg_color="gray", 
            hover_color="darkgray",
            height=28
        ).grid(row=0, column=1, padx=(10,0))
        
        self.refresh_ports()

        ctk.CTkButton(self, text="Connect", command=self.start_app, height=35, fg_color="#4EB9D7", hover_color="#357D92").pack(padx=20, pady=(10,20), fill="x", expand=True)

    def refresh_ports(self):
        ports = [port.device for port in serial.tools.list_ports.comports()]
        if ports:
            self.port_combo.configure(values=ports)
            self.selected_port.set(ports[0])
        else:
            self.port_combo.configure(values=["No ports found"])
            self.selected_port.set("No ports found")

    def start_app(self):
        chosen_port = self.selected_port.get()
        if "No ports" in chosen_port:
            messagebox.showerror("Error", "No COM port found. Check device and refresh.")
            return
        self.destroy()
        app = RecyclingApp(selected_port=chosen_port)
        app.mainloop()

if __name__ == "__main__":
    ctk.set_appearance_mode("Light")
    ctk.set_default_color_theme("dark-blue")
    
    # Create dummy logo for demo if missing
    if not os.path.exists(LOGO_PATH):
        try:
            os.makedirs("assets", exist_ok=True) # Ensure folder exists
            img = Image.new('RGB', LOGO_SIZE, color = 'gray')
            d = ImageDraw.Draw(img)
            d.text((10,20), "ASORATE", fill=(0,0,0))
            # Saves to root if folder fails, just to be safe
            img.save(LOGO_PATH) 
        except Exception: pass

    COMPortSelection().mainloop()